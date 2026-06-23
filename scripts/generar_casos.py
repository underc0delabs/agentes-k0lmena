#!/usr/bin/env python3
"""Genera los casos de prueba a partir de un JSON: la planilla Excel (.xlsx) y,
en el mismo paso, una versión en Markdown (.md) con los casos en tabla ASCII alineada.

Uso:
    python scripts/generar_casos.py <entrada.json> <salida.xlsx> [--limpiar]

Escribe <salida.xlsx> y, al lado, <salida.md> (misma ruta, extensión .md).
Con --limpiar, borra el JSON de entrada al terminar (útil para los temporales).

Formato del JSON de entrada:
{
  "proyecto": "NOMBRE DEL PROYECTO",
  "modulo": "HU-001 — Inicio de sesión",
  "casos": [
    {"id": "CP-001", "titulo": "...", "descripcion": "...", "precondiciones": "...",
     "datos": "Email: ...\nContraseña: ...", "pasos": "1. ...\n2. ...", "resultado": "...",
     "estado": "Pendiente", "prioridad": "Alta", "etiquetas": "@login",
     "evidencia": "", "fecha_ejecucion": "", "comentarios": ""}
  ]
}
"""
import sys
import os
import re
import json
import textwrap
import subprocess

COLS = ["ID #", "Título", "Descripción", "Precondiciones", "Datos de Prueba", "Pasos",
        "Resultado Esperado", "Estado", "Prioridad", "#Etiquetas", "Evidencia",
        "Fecha de Ejecución", "Comentarios"]
KEYS = ["id", "titulo", "descripcion", "precondiciones", "datos", "pasos", "resultado",
        "estado", "prioridad", "etiquetas", "evidencia", "fecha_ejecucion", "comentarios"]
# El .md tiene: (1) una tabla Resumen y (2) la tabla con todos los casos (una fila por caso).
# Resumen: 4 columnas, ordenado por prioridad (lo más crítico primero).
RESUMEN_COLS = ["ID", "Título", "Resultado esperado", "Prioridad"]
RESUMEN_KEYS = ["id", "titulo", "resultado", "prioridad"]
RESUMEN_W = [6, 24, 32, 10]

# Tabla de casos del .md. Quita Estado (siempre "Pendiente"), Evidencia, Fecha de Ejecución
# y Comentarios para no ensanchar de más; todas esas columnas sí están en el .xlsx.
MD_COLS = ["ID #", "Título", "Descripción", "Precondiciones", "Datos de Prueba", "Pasos",
           "Resultado Esperado", "Prioridad", "#Etiquetas"]
MD_KEYS = ["id", "titulo", "descripcion", "precondiciones", "datos", "pasos",
           "resultado", "prioridad", "etiquetas"]
# Ancho máximo por columna (las celdas largas se ajustan a varias líneas dentro de la celda).
MD_MAXW = [6, 16, 22, 16, 18, 24, 20, 9, 12]

# Orden de prioridad: más crítico primero, prioridad baja al final.
PRIORIDAD_ORDEN = {"crítica": 0, "critica": 0, "alta": 1, "media": 2, "baja": 3}


def _need(pkg, instalar=None):
    """Asegura que la librería esté disponible. Si falta, intenta instalarla; si no
    puede, avisa con un mensaje claro y corta (en vez de fallar con un traceback)."""
    instalar = instalar or pkg
    try:
        __import__(pkg)
        return
    except ImportError:
        pass
    subprocess.run([sys.executable, "-m", "pip", "install", instalar, "--quiet"], check=False)
    try:
        __import__(pkg)
    except ImportError:
        sys.exit(
            f"\n[!] Falta la librería '{instalar}' y no se pudo instalar sola.\n"
            f"    Instalá las dependencias con:  pip install -r requirements.txt\n"
            f"    (o directamente:  pip install {instalar})\n")


def to_xlsx(data, out):
    _need("openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    proyecto = data.get("proyecto", "NOMBRE DEL PROYECTO")
    modulo = data.get("modulo", "Módulo")
    casos = data.get("casos", [])
    widths = [6, 30, 42, 32, 30, 46, 40, 13, 11, 22, 20, 18, 30]
    n = len(COLS)
    last_col = chr(ord("A") + n - 1)

    wb = Workbook()
    ws = wb.active
    ws.title = (re.sub(r"[:\\/?*\[\]]", "-", modulo)[:31]) or "Casos"
    gris = PatternFill("solid", fgColor="404040")
    azul = PatternFill("solid", fgColor="4472C4")
    blanco_bold = Font(name="Arial", color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = proyecto
    ws["A1"].fill = gris
    ws["A1"].font = Font(name="Arial", color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Módulo: {modulo}"
    ws["A2"].fill = azul
    ws["A2"].font = blanco_bold
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    for i, c in enumerate(COLS):
        cell = ws.cell(row=3, column=i + 1, value=c)
        cell.fill = azul
        cell.font = blanco_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[3].height = 24
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w

    data_font = Font(name="Arial", size=10)
    start = 4
    for r, caso in enumerate(casos):
        row = start + r
        for c, key in enumerate(KEYS):
            cell = ws.cell(row=row, column=c + 1, value=caso.get(key, ""))
            cell.font = data_font
            cell.border = borde
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="center" if c in (0, 7, 8) else "left")
        ws.row_dimensions[row].height = 95

    last_row = start + len(casos) + 50
    dv_estado = DataValidation(type="list",
        formula1='"N/A,Pendiente,En ejecución,Aprobado,Fallido,Bloqueado"', allow_blank=True)
    dv_prio = DataValidation(type="list", formula1='"Crítica,Alta,Media,Baja"', allow_blank=True)
    ws.add_data_validation(dv_estado)
    ws.add_data_validation(dv_prio)
    dv_estado.add(f"H{start}:H{last_row}")
    dv_prio.add(f"I{start}:I{last_row}")
    ws.freeze_panes = "A4"
    wb.save(out)


def _wrap(texto, ancho):
    """Ajusta el texto a 'ancho' caracteres, respetando los saltos de línea existentes."""
    lineas = []
    for linea in str(texto).split("\n"):
        if linea.strip() == "":
            lineas.append("")
        else:
            lineas.extend(textwrap.wrap(linea, width=ancho) or [""])
    return "\n".join(lineas)


def to_md(data, out):
    _need("tabulate")
    from tabulate import tabulate
    casos = data.get("casos", [])
    titulo = (f"Casos de prueba — {data.get('modulo', '')}").strip(" —")
    partes = [f"# {titulo}", "", "## Resumen", ""]
    filas_r = [[_wrap(c.get(k, ""), w) for k, w in zip(RESUMEN_KEYS, RESUMEN_W)] for c in casos]
    partes += ["```", tabulate(filas_r, headers=RESUMEN_COLS, tablefmt="grid"), "```", ""]
    partes += ["## Detalle de los casos", ""]
    filas = [[c.get(k, "") for k in MD_KEYS] for c in casos]
    tabla = tabulate(filas, headers=MD_COLS, tablefmt="grid", maxcolwidths=MD_MAXW)
    partes += ["```", tabla, "```", ""]
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(partes) + "\n")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    if len(args) < 2:
        print("Uso: python scripts/generar_casos.py <entrada.json> <salida.xlsx> [--limpiar]")
        sys.exit(1)
    entrada, salida = args[0], args[1]
    if not salida.lower().endswith(".xlsx"):
        print("La salida debe terminar en .xlsx")
        sys.exit(1)
    with open(entrada, encoding="utf-8") as f:
        data = json.load(f)
    # Ordenar los casos por prioridad: los más críticos primero, los de prioridad baja al final.
    # El orden es estable: dentro de la misma prioridad, se respeta el orden original.
    data["casos"] = sorted(
        data.get("casos", []),
        key=lambda c: PRIORIDAD_ORDEN.get(str(c.get("prioridad", "")).strip().lower(), 99))
    to_xlsx(data, salida)
    print(f"OK: {salida}")
    md = re.sub(r"\.xlsx$", ".md", salida, flags=re.IGNORECASE)
    to_md(data, md)
    print(f"OK: {md}")
    # Con --limpiar, borra el JSON de entrada (cross-platform, sin depender de 'rm').
    if "--limpiar" in flags:
        try:
            os.remove(entrada)
            print(f"Borrado temporal: {entrada}")
        except OSError:
            pass


if __name__ == "__main__":
    main()
